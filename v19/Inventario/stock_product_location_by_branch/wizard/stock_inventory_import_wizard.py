import base64
import csv
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_HEADER_KEYWORDS = {
    'codigo', 'código', 'producto', 'ubicacion', 'ubicación',
    'cantidad', 'code', 'product', 'location', 'qty', 'quantity',
    'nombre', 'name', 'articulo', 'artículo',
}


class StockInventoryImportWizard(models.TransientModel):
    _name = 'stock.inventory.import.wizard'
    _description = 'Importación Masiva de Inventario desde CSV/Excel'

    # ── Configuración ────────────────────────────────────────────────────────────

    template_file = fields.Binary(string='Plantilla', attachment=False, readonly=True)
    template_filename = fields.Char(default='plantilla_inventario.xlsx')

    company_id = fields.Many2one(
        'res.company', string='Empresa', required=True,
        default=lambda self: self.env.company,
    )
    warehouse_id = fields.Many2one(
        'stock.warehouse', string='Almacén', required=True,
        domain="[('company_id', '=', company_id)]",
    )
    import_file = fields.Binary(string='Archivo CSV o Excel', attachment=False)
    import_filename = fields.Char(string='Nombre del archivo')

    # ── Estado y resultados ──────────────────────────────────────────────────────

    state = fields.Selection([
        ('draft', 'Pendiente'),
        ('done', 'Procesado'),
    ], default='draft')

    result_new_configs = fields.Integer(string='Configuraciones nuevas')
    result_existing_configs = fields.Integer(string='Configuraciones ya existentes')
    result_new_quants = fields.Integer(string='Quants nuevos creados')
    result_updated_quants = fields.Integer(string='Quants actualizados')
    result_errors = fields.Text(string='Errores / Advertencias')
    result_quant_ids = fields.Char(string='IDs de quants procesados')  # comma-separated

    # ── Parseo de archivo ────────────────────────────────────────────────────────

    def _parse_file(self):
        """
        Devuelve lista de tuplas (producto_str, ubicacion_str, cantidad_float)
        a partir del archivo CSV o Excel cargado.
        Omite la primera fila si parece un encabezado.
        """
        if not self.import_file:
            raise UserError(_('Por favor seleccione un archivo.'))

        raw = base64.b64decode(self.import_file)
        fname = (self.import_filename or '').lower()

        rows = []
        if fname.endswith('.xlsx'):
            try:
                import openpyxl  # noqa: PLC0415
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
                for row in wb.active.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows.append([str(c).strip() if c is not None else '' for c in row])
            except ImportError:
                raise UserError(_('openpyxl no está disponible. Guarde el archivo como CSV y vuelva a intentarlo.'))
        else:
            try:
                text = raw.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = raw.decode('latin-1')
            first = text.split('\n')[0]
            sep = ';' if first.count(';') >= first.count(',') else ','
            for row in csv.reader(io.StringIO(text), delimiter=sep):
                if any(cell.strip() for cell in row):
                    rows.append([c.strip() for c in row])

        if not rows:
            raise UserError(_('El archivo está vacío.'))

        # Detectar y saltar encabezado
        if self._is_header(rows[0]):
            rows = rows[1:]

        return rows

    @staticmethod
    def _is_header(row):
        return any(str(cell).strip().lower() in _HEADER_KEYWORDS for cell in row if cell)

    # ── Búsqueda de registros ────────────────────────────────────────────────────

    def _find_product(self, value):
        value = value.strip()
        if not value:
            return None

        # Resolver external ID de Odoo (ej: __export__.product_template_96996_0633f7c9)
        if '.' in value:
            try:
                record = self.env.ref(value, raise_if_not_found=False)
                if record and record._name == 'product.template':
                    variant = record.sudo().product_variant_ids.filtered('is_storable')[:1]
                    if variant:
                        return variant
                elif record and record._name == 'product.product' and record.sudo().is_storable:
                    return record.sudo()
            except Exception:
                pass

        P = self.env['product.product'].sudo()
        return (
            P.search([('default_code', '=', value), ('is_storable', '=', True)], limit=1)
            or P.search([('name', '=', value), ('is_storable', '=', True)], limit=1)
            or None
        )

    def _find_location(self, value):
        value = value.strip()
        if not value:
            return None
        L = self.env['stock.location'].sudo()
        return (
            L.search([('complete_name', '=', value), ('usage', '=', 'internal'), ('active', '=', True)], limit=1)
            or L.search([('name', '=', value), ('usage', '=', 'internal'), ('active', '=', True)], limit=1)
            or None
        )

    # ── Acción principal ─────────────────────────────────────────────────────────

    def action_import(self):
        self.ensure_one()
        rows = self._parse_file()

        Quant = self.env['stock.quant'].sudo()
        Config = self.env['stock.product.branch.location'].sudo()

        new_configs = existing_configs = new_quants = updated_quants = 0
        errors = []
        quant_ids = []

        for i, row in enumerate(rows, start=2):
            if len(row) < 3:
                errors.append(_('Fila %(n)s: se esperan 3 columnas (Código, Ubicación, Cantidad).', n=i))
                continue

            product_val, location_val, qty_val = row[0], row[1], row[2]

            if not product_val and not location_val:
                continue

            # Cantidad
            try:
                quantity = float(str(qty_val).replace(',', '.'))
            except (ValueError, TypeError):
                errors.append(_('Fila %(n)s: cantidad inválida "%(v)s".', n=i, v=qty_val))
                continue

            # Producto
            product = self._find_product(product_val)
            if not product:
                errors.append(_('Fila %(n)s: producto no encontrado "%(v)s".', n=i, v=product_val))
                continue

            # Ubicación
            location = self._find_location(location_val)
            if not location:
                errors.append(_('Fila %(n)s: ubicación no encontrada "%(v)s".', n=i, v=location_val))
                continue

            # Config producto→almacén→ubicación
            cfg = Config.search([
                ('product_id', '=', product.id),
                ('warehouse_id', '=', self.warehouse_id.id),
                ('company_id', '=', self.company_id.id),
            ], limit=1)
            if not cfg:
                try:
                    Config.create({
                        'company_id': self.company_id.id,
                        'warehouse_id': self.warehouse_id.id,
                        'product_id': product.id,
                        'location_id': location.id,
                    })
                    new_configs += 1
                except Exception as e:
                    errors.append(_('Fila %(n)s: error al crear configuración — %(e)s', n=i, e=e))
                    continue
            else:
                existing_configs += 1

            # Quant
            quant = Quant.search([
                ('product_id', '=', product.id),
                ('location_id', '=', location.id),
            ], limit=1)
            try:
                if not quant:
                    quant = Quant.with_context(inventory_mode=True).create({
                        'product_id': product.id,
                        'location_id': location.id,
                        'inventory_quantity': quantity,
                    })
                    new_quants += 1
                else:
                    quant.write({'inventory_quantity': quantity})
                    updated_quants += 1
                quant_ids.append(quant.id)
            except Exception as e:
                errors.append(_('Fila %(n)s: error al crear/actualizar quant — %(e)s', n=i, e=e))
                _logger.exception('SPLB Import: fila %s', i)

        self.write({
            'state': 'done',
            'result_new_configs': new_configs,
            'result_existing_configs': existing_configs,
            'result_new_quants': new_quants,
            'result_updated_quants': updated_quants,
            'result_errors': '\n'.join(errors) if errors else False,
            'result_quant_ids': ','.join(str(x) for x in quant_ids) if quant_ids else False,
        })

        # Reabrir el mismo wizard en estado "done"
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_download_template(self):
        self.ensure_one()
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill, Alignment  # noqa: PLC0415

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        # Encabezados
        headers = ['Codigo', 'Ubicacion', 'Cantidad']
        header_fill = PatternFill(start_color='2F4F6F', end_color='2F4F6F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Filas de ejemplo
        examples = [
            ['00208087', 'DP-CO/Existencias/Pasillo 1 - Sector A', 0],
            ['KC2106HSB', 'DP-CO/Existencias/Pasillo 2 - Sector B', 0],
            ['SXR280',    'DP-CO/Existencias/Pasillo 1 - Sector A', 0],
        ]
        for row_idx, row_data in enumerate(examples, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        self.write({
            'template_file': base64.b64encode(buf.getvalue()),
            'template_filename': 'plantilla_inventario.xlsx',
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_open_inventory(self):
        self.ensure_one()
        if not self.result_quant_ids:
            raise UserError(_('No hay quants procesados para mostrar.'))
        ids = [int(x) for x in self.result_quant_ids.split(',') if x.strip()]
        action = self.env['stock.quant'].action_view_inventory()
        action['domain'] = [('id', 'in', ids)]
        action['name'] = _('Inventario Físico — Importación Masiva')
        action['target'] = 'main'
        return action
